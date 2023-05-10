import os
import csv
import openpyxl


def process_csv_files():
    current_directory = os.getcwd()
    csv_files = [f for f in os.listdir(current_directory) if f.endswith('.csv')]
    master_workbook_path = os.path.join(current_directory, 'master.xlsm')
    wb = openpyxl.load_workbook(master_workbook_path, read_only=False)
    ws = wb.active

    def find_data_in_master(domainname_raw):

        for row in ws.iter_rows():
            if row[0].value == domainname_raw:
                alerting = row[ws['GM'][0].column - 1].value
                site_raw = row[ws['GL'][0].column - 1].value
                try:
                    site = "YAL" + "".join(filter(str.isdigit, site_raw))
                except:
                    site = site_input
                return alerting, site
        return None

    for csv_file in csv_files:
        output_filename = os.path.splitext(csv_file)[0] + "-BAT.csv"

        site_input = input(f"What site is this? ({csv_file}): ")
        partition = site_input + "-Staging-PT"
        location = site_input + "-LOC"

        with open(csv_file, 'r') as input_file, open(output_filename, 'w', newline='') as output_file:
            csv_reader = csv.reader(input_file)
            csv_writer = csv.writer(output_file)

            header_line = ["DOMAIN NAME", "SLOT", "SUBUNIT", "PORT NUMBER", "PORT DIRECTORY NUMBER", "Route Partition 1", "PORT DESCRIPTION", "CSS", "LOCATION", "LINE DESCRIPTION  1", "LINE CSS  1", "EXTERNAL PHONE NUMBER MASK  1", "ALERTING NAME  1"]
            csv_writer.writerow(header_line)

            for row in csv_reader:
                slot = row[0]
                subslot = row[1]
                port = row[2]

                domainname_raw = row[3]
                domainname = "SKIGW" + domainname_raw[2:-3]

                try:
                    dirnum = row[7]
                    if dirnum == "CCM":
                        continue
                except IndexError:
                    dirnum = "TBD"
                    print(f"IndexError in file: {csv_file}")

                dirnumfull = "20343" + dirnum
                alerting, site = find_data_in_master(domainname_raw)

                if site:
                    linedesc = f"{site} {dirnum}"
                else:
                    linedesc = ""

                portdesc = f"{linedesc} / {os.path.splitext(csv_file)[0]} {slot}-{subslot}-{port}"

                new_line = [domainname, slot, subslot, port, dirnum, partition, portdesc, "International-CSS", location,
                            linedesc, "International-CSS", dirnumfull, alerting]
                csv_writer.writerow(new_line)


process_csv_files()
