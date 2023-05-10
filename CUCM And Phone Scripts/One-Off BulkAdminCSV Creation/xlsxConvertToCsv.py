import csv
import openpyxl
import os
from openpyxl import load_workbook


def xlsx_to_csv(xlsx_file, csv_file):
    wb = load_workbook(xlsx_file)
    ws = wb.active

    with open(csv_file, 'w', newline='', encoding='utf-8') as f:
        csv_writer = csv.writer(f)

        for row in ws.iter_rows():
            row_data = [cell.value for cell in row]
            csv_writer.writerow(row_data)


def process_files_in_directory():
    script_directory = os.path.dirname(os.path.realpath(__file__))
    files = os.listdir(script_directory)

    for file in files:
        if file.endswith(".xlsx"):
            xlsx_file = os.path.join(script_directory, file)
            csv_file = os.path.join(script_directory, f"{os.path.splitext(file)[0]}.csv")
            xlsx_to_csv(xlsx_file, csv_file)
            print(f"Converted {xlsx_file} to {csv_file}")


if __name__ == "__main__":
    process_files_in_directory()